from __future__ import annotations

import csv
import json
import re
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

from openpyxl import load_workbook


REQUIRED_SHEETS = ("league_seasons", "managers", "budgets", "draft_results", "rankings")
SALE_REQUIRED_FIELDS = (
    "season",
    "overall_order",
    "player_name",
    "position",
    "winning_manager_id",
    "winning_price",
)


@dataclass(frozen=True)
class DatasetIssue:
    severity: str
    code: str
    message: str
    sheet: str = ""
    row: Optional[int] = None


@dataclass(frozen=True)
class DatasetBuildReport:
    ready_for_price_training: bool
    source_workbook: str
    output_directory: str
    auction_sale_count: int
    keeper_count: int
    college_promotion_count: int
    ranking_count: int
    issues: Tuple[DatasetIssue, ...]

    def to_dict(self) -> dict:
        payload = asdict(self)
        payload["issues"] = [asdict(issue) for issue in self.issues]
        return payload


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _integer(value: object) -> Optional[int]:
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _number(value: object) -> Optional[float]:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def canonical_league_key(value: object) -> str:
    text = _text(value).lower()
    if "bishop" in text:
        return "bishop_sycamore"
    if "gdfm" in text or "gdpfm" in text or "gpdfm" in text:
        return "gdfm"
    return "_".join(text.replace("-", " ").split())


def _sheet_rows(worksheet) -> List[Dict[str, object]]:
    values = list(worksheet.iter_rows(values_only=True))
    if not values:
        return []
    headers = [_text(value) for value in values[0]]
    rows = []
    for row_number, values_row in enumerate(values[1:], start=2):
        row = {
            header: values_row[index] if index < len(values_row) else None
            for index, header in enumerate(headers)
            if header
        }
        row["_source_row"] = row_number
        rows.append(row)
    return rows


def _truthy(value: object) -> bool:
    if isinstance(value, bool):
        return value
    return _text(value).lower() in {"1", "true", "yes", "y"}


def _extract_note_number(notes: object, label: str) -> Optional[float]:
    pattern = r"{0}:\s*(-?\d+(?:\.\d+)?)".format(re.escape(label))
    match = re.search(pattern, _text(notes), flags=re.IGNORECASE)
    return float(match.group(1)) if match else None


def _retained_phase_cutoffs(rows: Sequence[Mapping[str, object]]) -> Dict[Tuple[str, int], int]:
    cutoffs: Dict[Tuple[str, int], int] = {}
    for row in rows:
        league_key = canonical_league_key(row.get("league_key"))
        season = _integer(row.get("season"))
        order = _integer(row.get("overall_order"))
        if league_key != "bishop_sycamore" or season is None or order is None:
            continue
        if _truthy(row.get("was_keeper")):
            key = (league_key, season)
            cutoffs[key] = max(cutoffs.get(key, 0), order)
    return cutoffs


def _record_type(
    row: Mapping[str, object],
    cutoffs: Mapping[Tuple[str, int], int],
) -> str:
    league_key = canonical_league_key(row.get("league_key"))
    season = _integer(row.get("season")) or 0
    order = _integer(row.get("overall_order")) or 0
    if league_key != "bishop_sycamore":
        return "auction_sale"
    cutoff = int(cutoffs.get((league_key, season), 0))
    if order <= cutoff:
        return "keeper" if _truthy(row.get("was_keeper")) else "college_promotion"
    return "auction_sale"


def _manager_resolution(
    manager_rows: Sequence[Mapping[str, object]],
) -> Tuple[Dict[Tuple[str, int, str], str], Dict[Tuple[str, int, str], str]]:
    by_user = {}
    by_alias = {}
    for row in manager_rows:
        league_key = canonical_league_key(row.get("league_key"))
        season = _integer(row.get("season")) or 0
        manager_id = _text(_literal(row.get("manager_id")))
        if not manager_id:
            continue
        user_id = _text(_literal(row.get("sleeper_user_id")))
        if user_id:
            by_user[(league_key, season, user_id)] = manager_id
        for field_name in ("manager_name", "Alias1", "Alias2"):
            alias = normalize_label(row.get(field_name))
            if alias:
                by_alias[(league_key, season, alias)] = manager_id
    return by_user, by_alias


def normalize_label(value: object) -> str:
    return "".join(character for character in _text(value).lower() if character.isalnum())


def _canonical_draft_rows(
    rows: Sequence[Mapping[str, object]],
    manager_rows: Sequence[Mapping[str, object]],
) -> List[dict]:
    cutoffs = _retained_phase_cutoffs(rows)
    managers_by_user, managers_by_alias = _manager_resolution(manager_rows)
    results = []
    for row in rows:
        league_key = canonical_league_key(row.get("league_key"))
        season = _integer(row.get("season")) or 0
        manager_id = _text(_literal(row.get("managerID")))
        if not manager_id:
            manager_id = managers_by_user.get(
                (league_key, season, _text(row.get("winning_manager_id"))),
                "",
            )
        if not manager_id:
            manager_id = managers_by_alias.get(
                (league_key, season, normalize_label(row.get("winning_manager_name"))),
                "",
            )
        results.append(
            {
                "league_key": league_key,
                "season": season,
                "draft_id": _text(row.get("draft_id")),
                "overall_order": _integer(row.get("overall_order")),
                "nomination_number": _integer(row.get("nomination_number")),
                "player_name": _text(row.get("player_name")),
                "sleeper_player_id": _text(row.get("sleeper_player_id")),
                "position": _text(row.get("position")).upper(),
                "winning_roster_id": _text(row.get("winning_roster_id")),
                "winning_manager_id": manager_id,
                "winning_manager_name": _text(row.get("winning_manager_name")),
                "winning_price": _integer(row.get("winning_price")),
                "record_type": _record_type(row, cutoffs),
                "modeled_market_value": _extract_note_number(
                    row.get("notes"), "App modeled market value"
                ),
                "app_do_not_exceed": _extract_note_number(
                    row.get("notes"), "app do-not-exceed"
                ),
                "notes": _text(row.get("notes")),
                "source_row": _integer(row.get("_source_row")),
            }
        )
    return results


def _canonical_rankings(rows: Sequence[Mapping[str, object]]) -> List[dict]:
    results = []
    for row in rows:
        results.append(
            {
                "ranking_source": _text(
                    row.get("ranking_source") or row.get("ranking_sourc")
                ),
                "ranking_season": _integer(row.get("ranking_season")),
                "ranking_date": _text(row.get("ranking_date")),
                "ranking_type": _text(row.get("ranking_type")),
                "scoring_format": _text(row.get("scoring_format")),
                "overall_rank": _integer(row.get("overall_rank")),
                "position_rank": _text(row.get("position_rank")),
                "player_name": _text(row.get("player_name")),
                "sleeper_player_id": _text(row.get("sleeper_player_id")),
                "position": _text(row.get("position")).upper(),
                "projected_points": _number(row.get("projected_points")),
                "auction_value": _number(row.get("auction_value")),
                "bye_week": _integer(row.get("bye_week")),
                "team": _text(row.get("team")),
                "notes": _text(row.get("notes")),
                "source_row": _integer(row.get("_source_row")),
            }
        )
    return results


def _literal(value: object) -> object:
    if isinstance(value, str) and value.startswith("="):
        return None
    return value


def _canonical_managers(rows: Sequence[Mapping[str, object]]) -> List[dict]:
    return [
        {
            "league_key": canonical_league_key(row.get("league_key")),
            "season": _integer(row.get("season")),
            "manager_id": _text(_literal(row.get("manager_id"))),
            "manager_name": _text(row.get("manager_name")),
            "sleeper_user_id": _text(_literal(row.get("sleeper_user_id"))),
            "alias_1": _text(row.get("Alias1")),
            "alias_2": _text(row.get("Alias2")),
            "source_row": _integer(row.get("_source_row")),
        }
        for row in rows
    ]


def _canonical_budgets(rows: Sequence[Mapping[str, object]]) -> List[dict]:
    def number(row: Mapping[str, object], field_name: str) -> Optional[float]:
        return _number(_literal(row.get(field_name)))

    return [
        {
            "league_key": canonical_league_key(row.get("league_key")),
            "season": _integer(row.get("season")),
            "manager_id": _text(_literal(row.get("manager_id"))),
            "sleeper_user_id": _text(_literal(row.get("sleeper_user_id"))),
            "base_budget": number(row, "base_budget"),
            "keeper_spend": number(row, "keeper_spend"),
            "open_auction_budget": number(row, "open_auction_budget"),
            "open_roster_spots": _integer(_literal(row.get("open_roster_spots"))),
            "notes": _text(row.get("notes")),
            "source_row": _integer(row.get("_source_row")),
        }
        for row in rows
    ]


def _canonical_league_seasons(rows: Sequence[Mapping[str, object]]) -> List[dict]:
    results = []
    for row in rows:
        result = {
            key: _literal(value)
            for key, value in row.items()
            if key != "_source_row"
        }
        result["league_key"] = canonical_league_key(row.get("league_key"))
        result["season"] = _integer(row.get("season"))
        result["sleeper_league_id"] = _text(_literal(row.get("sleeper_league_id")))
        result["sleeper_draft_id"] = _text(_literal(row.get("sleeper_draft_id")))
        result["source_row"] = _integer(row.get("_source_row"))
        results.append(result)
    return results


def _formula_issues(workbook, issues: List[DatasetIssue]) -> None:
    for sheet_name in ("budgets", "managers"):
        worksheet = workbook[sheet_name]
        formula_rows = set()
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_rows.add(cell.row)
                    break
        if formula_rows:
            issues.append(
                DatasetIssue(
                    severity="warning",
                    code="formula_input_not_canonical",
                    message=(
                        "{0} rows contain formulas and are excluded from canonical "
                        "price inputs until explicit values are supplied."
                    ).format(len(formula_rows)),
                    sheet=sheet_name,
                )
            )


def _validate_sales(rows: Sequence[Mapping[str, object]], issues: List[DatasetIssue]) -> None:
    seen = set()
    for row in rows:
        if row["record_type"] != "auction_sale":
            continue
        for field_name in SALE_REQUIRED_FIELDS:
            if row.get(field_name) in (None, ""):
                issues.append(
                    DatasetIssue(
                        severity="error",
                        code="missing_sale_field",
                        message="Auction sale is missing {0}.".format(field_name),
                        sheet="draft_results",
                        row=row.get("source_row"),
                    )
                )
        key = (row.get("league_key"), row.get("season"), row.get("overall_order"))
        if key in seen:
            issues.append(
                DatasetIssue(
                    severity="error",
                    code="duplicate_sale_order",
                    message="Duplicate league/season sale order {0}.".format(key),
                    sheet="draft_results",
                    row=row.get("source_row"),
                )
            )
        seen.add(key)


def _validate_rankings(rows: Sequence[Mapping[str, object]], issues: List[DatasetIssue]) -> None:
    seen = set()
    for row in rows:
        key = (
            row.get("ranking_source"),
            row.get("ranking_season"),
            row.get("overall_rank"),
            row.get("player_name"),
        )
        if key in seen:
            issues.append(
                DatasetIssue(
                    severity="error",
                    code="duplicate_ranking",
                    message="Duplicate ranking row {0}.".format(key),
                    sheet="rankings",
                    row=row.get("source_row"),
                )
            )
        seen.add(key)


def _write_csv(path: Path, rows: Sequence[Mapping[str, object]]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def build_canonical_history_dataset(
    workbook_path: Path,
    output_directory: Path,
) -> DatasetBuildReport:
    workbook_path = Path(workbook_path)
    output_directory = Path(output_directory)
    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    missing = [name for name in REQUIRED_SHEETS if name not in workbook.sheetnames]
    if missing:
        raise ValueError("Workbook is missing required sheets: {0}".format(", ".join(missing)))

    issues: List[DatasetIssue] = []
    raw_manager_rows = _sheet_rows(workbook["managers"])
    draft_rows = _canonical_draft_rows(
        _sheet_rows(workbook["draft_results"]),
        raw_manager_rows,
    )
    rankings = _canonical_rankings(_sheet_rows(workbook["rankings"]))
    managers = _canonical_managers(raw_manager_rows)
    budgets = _canonical_budgets(_sheet_rows(workbook["budgets"]))
    league_seasons = _canonical_league_seasons(
        _sheet_rows(workbook["league_seasons"])
    )
    _validate_sales(draft_rows, issues)
    _validate_rankings(rankings, issues)
    _formula_issues(workbook, issues)

    sales = [row for row in draft_rows if row["record_type"] == "auction_sale"]
    keepers = [row for row in draft_rows if row["record_type"] == "keeper"]
    promotions = [
        row for row in draft_rows if row["record_type"] == "college_promotion"
    ]

    if any(row["league_key"] == "gdfm" for row in sales):
        issues.append(
            DatasetIssue(
                severity="warning",
                code="gdfm_incomplete_end_state",
                message=(
                    "GDFM has eight known unrecorded roster slots; observed sales are "
                    "valid for price training, but ending-cash labels are unavailable."
                ),
                sheet="draft_results",
            )
        )
    issues.append(
        DatasetIssue(
            severity="warning",
            code="ranking_snapshot_timing",
            message=(
                "Historical FantasyPros snapshots are near-draft proxies and may have "
                "been finalized after their auctions."
            ),
            sheet="rankings",
        )
    )

    output_directory.mkdir(parents=True, exist_ok=True)
    _write_csv(output_directory / "sales.csv", sales)
    _write_csv(output_directory / "keepers.csv", keepers)
    _write_csv(output_directory / "college_promotions.csv", promotions)
    _write_csv(output_directory / "rankings.csv", rankings)
    _write_csv(output_directory / "managers.csv", managers)
    _write_csv(output_directory / "budgets.csv", budgets)
    _write_csv(output_directory / "league_seasons.csv", league_seasons)

    report = DatasetBuildReport(
        ready_for_price_training=not any(issue.severity == "error" for issue in issues),
        source_workbook=str(workbook_path),
        output_directory=str(output_directory),
        auction_sale_count=len(sales),
        keeper_count=len(keepers),
        college_promotion_count=len(promotions),
        ranking_count=len(rankings),
        issues=tuple(issues),
    )
    (output_directory / "validation_report.json").write_text(
        json.dumps(report.to_dict(), indent=2, sort_keys=True),
        encoding="utf-8",
    )
    return report
