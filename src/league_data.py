from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook

from src.league_config import (
    DRAFT_START_MARKER,
    HISTORICAL_DRAFT_SHEETS,
    MANAGERS,
)


# =========================================================
# DATA OBJECTS
# =========================================================

@dataclass
class KeeperOption:
    player_name: str
    position: str
    keeper_cost: Optional[int]
    source_row: int


@dataclass
class ManagerLeagueData:
    manager_id: str
    spreadsheet_tab: str

    pre_keeper_budget: int

    keeper_options: List[KeeperOption] = field(
        default_factory=list
    )


@dataclass
class HistoricalAuctionSale:
    year: int
    player_name: str
    price: int

    manager_id: Optional[str] = None
    manager_raw: Optional[str] = None

    source_row: Optional[int] = None


@dataclass
class LeagueWorkbookData:
    managers: Dict[str, ManagerLeagueData]

    historical_sales: List[
        HistoricalAuctionSale
    ]

    warnings: List[str] = field(
        default_factory=list
    )


# =========================================================
# NAME MAPPING
# =========================================================

OWNER_ALIASES = {

    # TALLY
    "zach": "tallevast",
    "zach tallevast": "tallevast",
    "tallevast": "tallevast",
    "tally": "tallevast",

    # JAYLEN
    "jaylen": "jaylen",
    "jaylen henderson": "jaylen",

    # BRANDON
    "brandon": "brandon",
    "brandon c": "brandon",
    "brandon carr": "brandon",

    # ERNEST
    "ernest": "ernest",
    "ern": "ernest",

    # DEREK
    "derek": "derek",

    # TED
    "ted": "ted_d",
    "ted d": "ted_d",
    "ted drury": "ted_d",

    # COSEY
    "cosey": "josh_cosey",
    "josh cosey": "josh_cosey",

    # NOBS
    "nobs": "nobs",
    "ben nobs": "nobs",

    # AUTREY
    "fritz": "autrey",
    "fritz autrey": "autrey",
    "autrey": "autrey",

    # SETH
    "seth": "seth",
    "seth f": "seth",

    # TROY
    "troy": "troy_l",
    "troy l": "troy_l",

    # STEPHEN / PETE
    "stephen": "stephen_m",
    "stephen m": "stephen_m",
    "pete": "stephen_m",
}


def normalize_text(value) -> str:
    if value is None:
        return ""

    return str(value).strip()


def normalize_owner(value) -> str:
    return normalize_text(
        value
    ).lower()


def resolve_owner(value) -> Optional[str]:
    """
    Resolve the various names used throughout
    the spreadsheet to our permanent manager_id.
    """

    key = normalize_owner(value)

    return OWNER_ALIASES.get(key)


def normalize_position(value) -> str:
    value = normalize_text(value).upper()

    value = value.strip()

    if value in {
        "D/ST",
        "DST",
        "DEFENSE",
    }:
        return "DEF"

    return value


def numeric_value(value) -> Optional[int]:
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return int(value)

    try:
        return int(
            float(
                str(value)
                .replace("$", "")
                .replace(",", "")
                .strip()
            )
        )

    except (ValueError, TypeError):
        return None


# =========================================================
# MAIN LOADER
# =========================================================

class LeagueDataLoader:

    def __init__(
        self,
        workbook_path,
    ):
        self.workbook_path = Path(
            workbook_path
        )

        if not self.workbook_path.exists():
            raise FileNotFoundError(
                f"Workbook not found: "
                f"{self.workbook_path}"
            )

        self.warnings: List[str] = []

        self.workbook = load_workbook(
            self.workbook_path,
            data_only=True,
        )

    # =====================================================
    # PUBLIC LOAD METHOD
    # =====================================================

    def load(self) -> LeagueWorkbookData:

        managers = (
            self.load_managers()
        )

        historical_sales = (
            self.load_historical_sales()
        )

        return LeagueWorkbookData(
            managers=managers,
            historical_sales=historical_sales,
            warnings=self.warnings,
        )

    # =====================================================
    # MANAGER SHEETS
    # =====================================================

    def load_managers(
        self,
    ) -> Dict[str, ManagerLeagueData]:

        result = {}

        for manager_id, identity in (
            MANAGERS.items()
        ):

            tab_name = (
                identity.spreadsheet_tab
            )

            if (
                tab_name
                not in self.workbook.sheetnames
            ):
                self.warnings.append(
                    f"Missing manager sheet: "
                    f"{tab_name}"
                )
                continue

            ws = self.workbook[
                tab_name
            ]

            budget = (
                self._extract_manager_budget(
                    ws
                )
            )

            if budget is None:
                self.warnings.append(
                    f"No draft budget found "
                    f"for {tab_name}."
                )

                budget = 0

            keeper_options = (
                self._extract_keeper_options(
                    ws,
                    tab_name,
                )
            )

            result[
                manager_id
            ] = ManagerLeagueData(
                manager_id=manager_id,
                spreadsheet_tab=tab_name,
                pre_keeper_budget=budget,
                keeper_options=keeper_options,
            )

        return result

    def _extract_manager_budget(
        self,
        ws,
    ) -> Optional[int]:

        # -----------------------------------------------
        # Preferred structure:
        #
        # Draft Budget | 385
        # -----------------------------------------------

        budget = self._find_value_next_to_label(
            ws,
            "Draft Budget",
        )

        if budget is not None:
            return numeric_value(
                budget
            )

        # -----------------------------------------------
        # Most of the existing sheets actually have:
        #
        # Salary | 375
        # Draft Budget | [blank]
        #
        # That Salary number is the current budget
        # brought into the 2026 keeper process.
        # -----------------------------------------------

        salary_budget = (
            self._find_value_next_to_label(
                ws,
                "Salary",
            )
        )

        if salary_budget is not None:
            return numeric_value(
                salary_budget
            )

        return None

    def _find_value_next_to_label(
        self,
        ws,
        target_label: str,
    ):
        """
        Find a workbook label and return the
        adjacent value immediately to its right.

        This deliberately ignores labels where the
        adjacent cell is blank.
        """

        target = (
            target_label
            .strip()
            .lower()
        )

        for row in ws.iter_rows():

            for cell in row:

                value = normalize_text(
                    cell.value
                ).lower()

                if value != target:
                    continue

                next_column = (
                    cell.column + 1
                )

                if (
                    next_column
                    > ws.max_column
                ):
                    continue

                next_value = ws.cell(
                    row=cell.row,
                    column=next_column,
                ).value

                if next_value is not None:
                    return next_value

        return None

    def _extract_keeper_options(
        self,
        ws,
        tab_name: str,
    ) -> List[KeeperOption]:

        header_row = None

        # Find:
        # Player | Position | Salary

        for row_number in range(
            1,
            min(
                ws.max_row,
                20,
            ) + 1,
        ):

            player_header = (
                normalize_text(
                    ws.cell(
                        row_number,
                        1,
                    ).value
                ).lower()
            )

            position_header = (
                normalize_text(
                    ws.cell(
                        row_number,
                        2,
                    ).value
                ).lower()
            )

            salary_header = (
                normalize_text(
                    ws.cell(
                        row_number,
                        3,
                    ).value
                ).lower()
            )

            if (
                player_header == "player"
                and
                position_header == "position"
                and
                salary_header == "salary"
            ):
                header_row = (
                    row_number
                )

                break

        if header_row is None:

            self.warnings.append(
                f"Could not find player table "
                f"on {tab_name}."
            )

            return []

        options = []

        seen_players = set()

        for row_number in range(
            header_row + 1,
            ws.max_row + 1,
        ):

            player_name = normalize_text(
                ws.cell(
                    row_number,
                    1,
                ).value
            )

            position = normalize_position(
                ws.cell(
                    row_number,
                    2,
                ).value
            )

            salary = numeric_value(
                ws.cell(
                    row_number,
                    3,
                ).value
            )

            # Not a roster row.
            if not player_name:
                continue

            if not position:
                continue

            player_key = (
                player_name
                .lower()
                .strip()
            )

            # Example: duplicate player rows.
            # Do not silently double-count them.
            if player_key in seen_players:

                self.warnings.append(
                    f"Duplicate player on "
                    f"{tab_name}: "
                    f"{player_name}"
                )

                continue

            seen_players.add(
                player_key
            )

            if salary is None:

                self.warnings.append(
                    f"Missing keeper salary: "
                    f"{tab_name} / "
                    f"{player_name}"
                )

            options.append(
                KeeperOption(
                    player_name=player_name,
                    position=position,
                    keeper_cost=salary,
                    source_row=row_number,
                )
            )

        return options

    # =====================================================
    # HISTORICAL AUCTION DATA
    # =====================================================

    def load_historical_sales(
        self,
    ) -> List[HistoricalAuctionSale]:

        sales = []

        for (
            year,
            sheet_name,
        ) in HISTORICAL_DRAFT_SHEETS.items():

            if (
                sheet_name
                not in self.workbook.sheetnames
            ):

                self.warnings.append(
                    f"Historical sheet missing: "
                    f"{sheet_name}"
                )

                continue

            ws = self.workbook[
                sheet_name
            ]

            start_row = (
                self._find_draft_start(
                    ws
                )
            )

            if start_row is None:

                # Important:
                # We DO NOT guess here because including
                # keepers as auction sales would poison
                # the model.
                self.warnings.append(
                    f"{sheet_name}: no "
                    f"'{DRAFT_START_MARKER}' "
                    f"marker found. "
                    f"Historical auction rows skipped."
                )

                continue

            year_sales = []

            for row_number in range(
                start_row,
                ws.max_row + 1,
            ):

                player_name = (
                    normalize_text(
                        ws.cell(
                            row_number,
                            1,
                        ).value
                    )
                )

                price = numeric_value(
                    ws.cell(
                        row_number,
                        2,
                    ).value
                )

                if (
                    not player_name
                    or price is None
                ):
                    continue

                raw_manager = (
                    normalize_text(
                        ws.cell(
                            row_number,
                            3,
                        ).value
                    )
                    or None
                )

                manager_id = None

                if raw_manager:

                    manager_id = (
                        resolve_owner(
                            raw_manager
                        )
                    )

                    if manager_id is None:

                        self.warnings.append(
                            f"{sheet_name} row "
                            f"{row_number}: "
                            f"unmapped manager "
                            f"'{raw_manager}'."
                        )

                sale = HistoricalAuctionSale(
                    year=year,
                    player_name=player_name,
                    price=price,
                    manager_id=manager_id,
                    manager_raw=raw_manager,
                    source_row=row_number,
                )

                year_sales.append(
                    sale
                )

            # Data-quality check
            if len(year_sales) < 20:

                self.warnings.append(
                    f"{sheet_name}: only "
                    f"{len(year_sales)} auction "
                    f"sales found after the "
                    f"draft marker."
                )

            sales.extend(
                year_sales
            )

        return sales

    def _find_draft_start(
        self,
        ws,
    ) -> Optional[int]:

        target = (
            DRAFT_START_MARKER
            .strip()
            .lower()
        )

        for row_number in range(
            1,
            ws.max_row + 1,
        ):

            value = normalize_text(
                ws.cell(
                    row_number,
                    1,
                ).value
            ).lower()

            if value == target:

                return (
                    row_number + 1
                )

        return None

