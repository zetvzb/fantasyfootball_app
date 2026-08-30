import csv
import json

from openpyxl import Workbook

from src.ml_history_dataset import (
    build_canonical_history_dataset,
    canonical_league_key,
)


def _workbook(path):
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in ("league_seasons", "budgets"):
        sheet = workbook.create_sheet(name)
        sheet.append(["league_key", "season", "value"])
        sheet.append(["Bishop Sycamore Alumni Support Group (2025)", 2025, "=1+1"])

    managers = workbook.create_sheet("managers")
    managers.append(
        ["league_key", "season", "manager_id", "manager_name",
         "sleeper_user_id", "Alias1", "Alias2", "manager_id"]
    )
    managers.append(
        ["Bishop Sycamore Alumni Support Group (2025)", 2025, "=A1", "Owner",
         "123456789012345", "owner", "Team", "bishop_owner"]
    )

    draft = workbook.create_sheet("draft_results")
    draft.append(
        [
            "league_key", "season", "draft_id", "overall_order",
            "nomination_number", "player_name", "sleeper_player_id",
            "position", "winning_roster_id", "winning_manager_id",
            "winning_manager_name", "winning_price", "was_keeper", "notes",
            "managerID",
        ]
    )
    draft.append(
        ["bishop_sycamore", 2025, "draft", 1, 1, "Keeper", "1", "WR", 1,
         "123456789012345", "Team", 20, True, "=VLOOKUP()", None]
    )
    draft.append(
        ["bishop_sycamore", 2025, "draft", 2, 2, "Promotion", "2", "RB", 1,
         "123456789012345", "Team", 10, False, "", None]
    )
    draft.append(
        ["bishop_sycamore", 2025, "draft", 3, 3, "Keeper Two", "5", "TE", 1,
         "123456789012345", "Team", 15, True, "", None]
    )
    draft.append(
        ["bishop_sycamore", 2025, "draft", 4, 4, "Auction", "3", "QB", 1,
         "123456789012345", "Team", 30, False, "", None]
    )
    draft.append(
        ["gdfm", 2026, "manual", 1, 1, "GDFM Sale", "4", "TE", 3,
         "team_3", "Troy", 12, False,
         "App modeled market value: 10.5; app do-not-exceed: 11", "team_3"]
    )

    rankings = workbook.create_sheet("rankings")
    rankings.append(
        [
            "ranking_source", "ranking_season", "ranking_date", "ranking_type",
            "scoring_format", "overall_rank", "position_rank", "player_name",
            "sleeper_player_id", "position", "projected_points", "auction_value",
            "bye_week", "team", "notes",
        ]
    )
    rankings.append(
        ["FantasyPros ECR", 2025, "2025-09-05", "draft", "half_ppr", 1,
         "WR1", "Ranked", "1", "WR", None, None, 5, "MIN", "proxy"]
    )
    workbook.save(path)


def test_canonical_league_key_handles_known_aliases():
    assert canonical_league_key("Bishop Sycamore Alumni Support Group (2025)") == "bishop_sycamore"
    assert canonical_league_key("manual_gdpfm_2026") == "gdfm"


def test_builder_classifies_retained_phase_and_exports_formula_free_tables(tmp_path):
    workbook_path = tmp_path / "history.xlsx"
    output = tmp_path / "canonical"
    _workbook(workbook_path)

    report = build_canonical_history_dataset(workbook_path, output)

    assert report.ready_for_price_training
    assert report.auction_sale_count == 2
    assert report.keeper_count == 2
    assert report.college_promotion_count == 1
    assert report.ranking_count == 1

    with (output / "sales.csv").open(encoding="utf-8") as file:
        sales = list(csv.DictReader(file))
    assert [sale["player_name"] for sale in sales] == ["Auction", "GDFM Sale"]
    assert sales[0]["winning_manager_id"] == "bishop_owner"
    assert sales[1]["modeled_market_value"] == "10.5"
    assert sales[1]["app_do_not_exceed"] == "11.0"
    assert (output / "managers.csv").exists()
    assert (output / "budgets.csv").exists()
    assert (output / "league_seasons.csv").exists()

    payload = json.loads((output / "validation_report.json").read_text())
    assert payload["ready_for_price_training"] is True
    assert any(
        issue["code"] == "formula_input_not_canonical"
        for issue in payload["issues"]
    )
